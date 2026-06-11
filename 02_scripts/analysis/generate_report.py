#!/usr/bin/env python3
"""
谷氨酸棒杆菌重测序分析报告生成器

读取流水线输出结果，生成包含以下内容的 Excel 报告：
  1. 样本变异汇总
  2. 高频变异详情
  3. 多态变异详情
  4. 样本间共有/特有变异对比
  5. snpEff 变异影响分类统计
  6. 基因水平影响详情

用法:
  python 02_scripts/generate_analysis_report.py [--out_dir OUTPUT_DIR]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── 路径配置 ─────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_TABLES_DIR = PROJECT_ROOT / "06_variant_calling" / "tables"
DEFAULT_ANNOT_DIR = PROJECT_ROOT / "06_variant_calling" / "annotated"
DEFAULT_OUTPUT    = PROJECT_ROOT / "06_variant_calling" / "analysis_report.xlsx"

SAMPLES = ["C0", "C32", "T0", "T32"]


# ── 样式定义 ─────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="微软雅黑", size=10)
TITLE_FONT  = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
SUB_FONT    = Font(name="微软雅黑", bold=True, size=11, color="2E75B6")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
SUM_FILL    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, max_col, max_width=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        lengths = []
        for cell in ws[letter]:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        best = max(lengths) + 3 if lengths else 10
        ws.column_dimensions[letter].width = min(best, max_width)


# ── 数据读取 ─────────────────────────────────────────────────────────────────

def read_table_tsv(path: Path) -> list[dict]:
    """读取 table.tsv 或 sites.tsv"""
    if not path.exists() or path.stat().st_size == 0:
        return []
    with open(path, "r") as f:
        reader = csv.DictReader(f, delimiter="\t")
        return list(reader)


def parse_variant_stats(path: Path) -> dict:
    """从 bcftools stats 文件提取关键数字"""
    result = {
        "records": 0, "snps": 0, "indels": 0, "mnps": 0,
        "ts": 0, "tv": 0, "ts_tv": 0.0,
        "substitutions": {},  # e.g. {"A>G": 3, ...}
    }
    if not path.exists():
        return result
    with open(path) as f:
        for line in f:
            if line.startswith("SN"):
                parts = line.strip().split("\t")
                key = parts[2]
                val = parts[3]
                if key == "number of records:":
                    result["records"] = int(val)
                elif key == "number of SNPs:":
                    result["snps"] = int(val)
                elif key == "number of indels:":
                    result["indels"] = int(val)
                elif key == "number of MNPs:":
                    result["mnps"] = int(val)
            elif line.startswith("TSTV"):
                parts = line.strip().split("\t")
                if len(parts) >= 6:
                    result["ts"] = int(parts[2])
                    result["tv"] = int(parts[3])
                    result["ts_tv"] = float(parts[4])
            elif line.startswith("ST"):
                parts = line.strip().split("\t")
                if len(parts) >= 4:
                    result["substitutions"][parts[2]] = int(parts[3])
    return result


def parse_snpeff_genes(path: Path) -> list[dict]:
    """读取 snpEff 的 genes.txt 文件"""
    if not path.exists():
        return []
    with open(path) as f:
        reader = csv.DictReader(f, delimiter="\t", skipinitialspace=True)
        return list(reader)


def count_genes_by_impact(genes: list[dict]) -> dict:
    """统计各影响等级的基因数"""
    high = moderate = low = modifier = 0
    high_genes = []
    moderate_genes = []
    for g in genes:
        h = int(g.get("variants_impact_HIGH", 0))
        m = int(g.get("variants_impact_MODERATE", 0))
        l = int(g.get("variants_impact_LOW", 0))
        mo = int(g.get("variants_impact_MODIFIER", 0))
        high += h
        moderate += m
        low += l
        modifier += mo
        if h > 0:
            high_genes.append(g.get("GeneName", ""))
        if m > 0:
            moderate_genes.append(g.get("GeneName", ""))
    return {"HIGH": high, "MODERATE": moderate, "LOW": low, "MODIFIER": modifier,
            "high_genes": high_genes, "moderate_genes": moderate_genes}


def get_sample_summary(sample: str, tables_dir: Path, annot_dir: Path) -> dict:
    """提取单个样本的汇总信息"""
    info = {"sample": sample}

    # 高频变异
    hf_table = tables_dir / f"{sample}.highfreq.table.tsv"
    hf_rows = read_table_tsv(hf_table)
    info["highfreq_count"] = len(hf_rows)

    hf_stats = annot_dir / "highfreq" / f"{sample}.highfreq.variant_stats.txt"
    hf_parsed = parse_variant_stats(hf_stats)
    info["highfreq_snps"] = hf_parsed["snps"]
    info["highfreq_indels"] = hf_parsed["indels"]
    info["highfreq_mnps"] = hf_parsed["mnps"]
    info["highfreq_ts"] = hf_parsed["ts"]
    info["highfreq_tv"] = hf_parsed["tv"]
    info["highfreq_ts_tv"] = hf_parsed["ts_tv"]

    if hf_rows:
        depths = [float(r.get("DP", 0) or 0) for r in hf_rows]
        info["highfreq_mean_dp"] = round(sum(depths) / len(depths), 1)
    else:
        info["highfreq_mean_dp"] = 0

    # 高频 snpEff 基因注释
    hf_genes_file = annot_dir / "highfreq" / f"{sample}.highfreq_snpEff_stats.genes.txt"
    hf_genes = parse_snpeff_genes(hf_genes_file)
    hf_impact = count_genes_by_impact(hf_genes)
    info["highfreq_HIGH"] = hf_impact["HIGH"]
    info["highfreq_MODERATE"] = hf_impact["MODERATE"]
    info["highfreq_LOW"] = hf_impact["LOW"]
    info["highfreq_MODIFIER"] = hf_impact["MODIFIER"]
    info["highfreq_high_genes"] = "; ".join(hf_impact["high_genes"])
    info["highfreq_moderate_genes"] = "; ".join(hf_impact["moderate_genes"])

    # 多态变异
    poly_table = tables_dir / f"{sample}.polymorphic.table.tsv"
    poly_rows = read_table_tsv(poly_table)
    info["poly_count"] = len(poly_rows)

    poly_stats = annot_dir / "polymorphic" / f"{sample}.polymorphic.variant_stats.txt"
    poly_parsed = parse_variant_stats(poly_stats)
    info["poly_snps"] = poly_parsed["snps"]
    info["poly_indels"] = poly_parsed["indels"]

    if poly_rows:
        depths = [float(r.get("DP", 0) or 0) for r in poly_rows]
        info["poly_mean_dp"] = round(sum(depths) / len(depths), 1)
    else:
        info["poly_mean_dp"] = 0

    # 多态 snpEff 基因注释
    poly_genes_file = annot_dir / "polymorphic" / f"{sample}.polymorphic_snpEff_stats.genes.txt"
    poly_genes = parse_snpeff_genes(poly_genes_file)
    poly_impact = count_genes_by_impact(poly_genes)
    info["poly_HIGH"] = poly_impact["HIGH"]
    info["poly_MODERATE"] = poly_impact["MODERATE"]
    info["poly_LOW"] = poly_impact["LOW"]
    info["poly_MODIFIER"] = poly_impact["MODIFIER"]

    return info


# ── Sheet 生成 ───────────────────────────────────────────────────────────────

def create_summary_sheet(wb, summaries: list[dict]):
    """Sheet 1: 样本变异汇总"""
    ws = wb.create_sheet("样本变异汇总", 0)

    # 标题
    ws.merge_cells("A1:L1")
    ws["A1"] = "谷氨酸棒杆菌重测序分析 - 样本变异汇总"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # 表头
    headers = [
        "样本", "高频变异数", "高频SNPs", "高频Indels", "高频Ts/Tv",
        "高频平均深度", "多态变异数", "多态SNPs", "多态Indels",
        "snpEff HIGH影响", "snpEff MODERATE影响",
        "snpEff HIGH影响的基因"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    for i, s in enumerate(summaries):
        row = 4 + i
        ws.cell(row=row, column=1, value=s["sample"])
        ws.cell(row=row, column=2, value=s["highfreq_count"])
        ws.cell(row=row, column=3, value=s["highfreq_snps"])
        ws.cell(row=row, column=4, value=s["highfreq_indels"])
        ws.cell(row=row, column=5, value=s["highfreq_ts_tv"])
        ws.cell(row=row, column=6, value=s["highfreq_mean_dp"])
        ws.cell(row=row, column=7, value=s["poly_count"])
        ws.cell(row=row, column=8, value=s["poly_snps"])
        ws.cell(row=row, column=9, value=s["poly_indels"])
        ws.cell(row=row, column=10, value=s["highfreq_HIGH"])
        ws.cell(row=row, column=11, value=s["highfreq_MODERATE"])
        ws.cell(row=row, column=12, value=s.get("highfreq_high_genes", ""))

    style_body(ws, 4, 3 + len(summaries), len(headers))
    auto_width(ws, len(headers))
    ws.sheet_properties.tabColor = "4472C4"


def create_highfreq_detail_sheet(wb, tables_dir: Path):
    """Sheet 2: 高频变异详情"""
    ws = wb.create_sheet("高频变异详情", 1)

    headers = ["CHROM", "POS", "REF", "ALT", "DP", "AO", "AF"]
    current_row = 1

    for sample in SAMPLES:
        # 样本标题
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        cell = ws.cell(row=current_row, column=1, value=f"▶ {sample} - 高频变异")
        cell.font = SUB_FONT
        current_row += 1

        # 表头
        for col, h in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=h)
        style_header(ws, current_row, len(headers))
        current_row += 1

        # 数据
        path = tables_dir / f"{sample}.highfreq.table.tsv"
        rows = read_table_tsv(path)
        if rows:
            for r in rows:
                ws.cell(row=current_row, column=1, value=r.get("CHROM", ""))
                ws.cell(row=current_row, column=2, value=int(r.get("POS", 0)))
                ws.cell(row=current_row, column=3, value=r.get("REF", ""))
                ws.cell(row=current_row, column=4, value=r.get("ALT", ""))
                ws.cell(row=current_row, column=5, value=int(r.get("DP", 0)))
                ws.cell(row=current_row, column=6, value=int(r.get("AO", 0)))
                ws.cell(row=current_row, column=7, value=round(float(r.get("AF", 0)), 4))
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value="(无)")
            current_row += 1

        current_row += 1  # 空行分隔

    # 样式
    style_body(ws, 2, current_row - 1, len(headers))
    auto_width(ws, len(headers))
    ws.sheet_properties.tabColor = "70AD47"


def create_poly_detail_sheet(wb, tables_dir: Path):
    """Sheet 3: 多态变异详情"""
    ws = wb.create_sheet("多态变异详情")

    headers = ["CHROM", "POS", "REF", "ALT", "DP", "AO", "AF"]
    current_row = 1

    for sample in SAMPLES:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        cell = ws.cell(row=current_row, column=1, value=f"▶ {sample} - 多态变异")
        cell.font = SUB_FONT
        current_row += 1

        for col, h in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=h)
        style_header(ws, current_row, len(headers))
        current_row += 1

        path = tables_dir / f"{sample}.polymorphic.table.tsv"
        rows = read_table_tsv(path)
        if rows:
            for r in rows:
                ws.cell(row=current_row, column=1, value=r.get("CHROM", ""))
                ws.cell(row=current_row, column=2, value=int(r.get("POS", 0)))
                ws.cell(row=current_row, column=3, value=r.get("REF", ""))
                ws.cell(row=current_row, column=4, value=r.get("ALT", ""))
                ws.cell(row=current_row, column=5, value=int(r.get("DP", 0)))
                ws.cell(row=current_row, column=6, value=int(r.get("AO", 0)))
                ws.cell(row=current_row, column=7, value=round(float(r.get("AF", 0)), 4))
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value="(无)")
            current_row += 1

        current_row += 1

    style_body(ws, 2, current_row - 1, len(headers))
    auto_width(ws, len(headers))
    ws.sheet_properties.tabColor = "FFC000"


def create_comparison_sheet(wb, tables_dir: Path):
    """Sheet 4: 样本间共有/特有变异"""
    ws = wb.create_sheet("变异对比分析")

    # 读取所有样本的高频位点
    sample_sites = {}
    for s in SAMPLES:
        path = tables_dir / f"{s}.highfreq.sites.tsv"
        sites = set()
        if path.exists() and path.stat().st_size > 0:
            with open(path) as f:
                for line in f:
                    line = line.strip()
                    if line:
                        sites.add(line)
        sample_sites[s] = sites

    # 共有位点
    all_sets = list(sample_sites.values())
    common = all_sets[0]
    for ss in all_sets[1:]:
        common &= ss

    # 生成表格
    ws.merge_cells("A1:D1")
    ws["A1"] = "样本间高频变异共有/特有统计"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    current_row = 3
    # 共有
    ws.cell(row=current_row, column=1, value="所有样本共有的高频变异位点")
    ws.cell(row=current_row, column=1).font = SUB_FONT
    current_row += 1
    ws.cell(row=current_row, column=1, value="CHROM")
    ws.cell(row=current_row, column=2, value="POS")
    style_header(ws, current_row, 2)
    current_row += 1
    if common:
        for site in sorted(common):
            chrom, pos = site.split("\t")
            ws.cell(row=current_row, column=1, value=chrom)
            ws.cell(row=current_row, column=2, value=int(pos))
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="(无)")
        current_row += 1
    current_row += 2

    # 特有
    ws.cell(row=current_row, column=1, value="各样本特有的高频变异位点")
    ws.cell(row=current_row, column=1).font = SUB_FONT
    current_row += 1

    for sample in SAMPLES:
        unique = sample_sites[sample] - common
        if unique:
            ws.cell(row=current_row, column=1, value=f"{sample} 特有 ({len(unique)} 个):")
            ws.cell(row=current_row, column=1).font = SUB_FONT
            current_row += 1
            ws.cell(row=current_row, column=1, value="CHROM")
            ws.cell(row=current_row, column=2, value="POS")
            style_header(ws, current_row, 2)
            current_row += 1
            for site in sorted(unique):
                chrom, pos = site.split("\t")
                ws.cell(row=current_row, column=1, value=chrom)
                ws.cell(row=current_row, column=2, value=int(pos))
                current_row += 1
            current_row += 1

    # 多态位点比较
    current_row += 1
    ws.cell(row=current_row, column=1, value="多态变异位点比较")
    ws.cell(row=current_row, column=1).font = SUB_FONT
    current_row += 1
    ws.cell(row=current_row, column=1, value="样本")
    ws.cell(row=current_row, column=2, value="多态位点数")
    ws.cell(row=current_row, column=3, value="位点位置")
    style_header(ws, current_row, 3)
    current_row += 1
    for sample in SAMPLES:
        path = tables_dir / f"{sample}.polymorphic.sites.tsv"
        poly_sites = []
        if path.exists() and path.stat().st_size > 0:
            with open(path) as f:
                for line in f:
                    line = line.strip()
                    if line:
                        poly_sites.append(line)
        ws.cell(row=current_row, column=1, value=sample)
        ws.cell(row=current_row, column=2, value=len(poly_sites))
        ws.cell(row=current_row, column=3, value="; ".join(poly_sites) if poly_sites else "无")
        current_row += 1

    style_body(ws, 3, current_row - 1, 3)
    auto_width(ws, 3)
    ws.sheet_properties.tabColor = "ED7D31"


def create_impact_sheet(wb, annot_dir: Path):
    """Sheet 5: snpEff 变异影响分类统计"""
    ws = wb.create_sheet("snpEff影响分类")

    ws.merge_cells("A1:F1")
    ws["A1"] = "snpEff 变异影响等级统计"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["样本", "类别", "HIGH", "MODERATE", "LOW", "MODIFIER"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    current_row = 4
    for sample in SAMPLES:
        for cat in ("highfreq", "polymorphic"):
            genes_file = annot_dir / cat / f"{sample}.{cat}_snpEff_stats.genes.txt"
            genes = parse_snpeff_genes(genes_file)
            impact = count_genes_by_impact(genes)

            ws.cell(row=current_row, column=1, value=sample)
            ws.cell(row=current_row, column=2, value=cat)
            ws.cell(row=current_row, column=3, value=impact["HIGH"])
            ws.cell(row=current_row, column=4, value=impact["MODERATE"])
            ws.cell(row=current_row, column=5, value=impact["LOW"])
            ws.cell(row=current_row, column=6, value=impact["MODIFIER"])
            current_row += 1

    style_body(ws, 4, current_row - 1, len(headers))
    auto_width(ws, len(headers))
    ws.sheet_properties.tabColor = "A42847"


def create_gene_detail_sheet(wb, annot_dir: Path):
    """Sheet 6: 基因水平影响详情"""
    ws = wb.create_sheet("基因影响详情")

    headers = [
        "样本", "类别", "基因名", "基因ID", "HIGH", "LOW",
        "MODERATE", "MODIFIER", "相关效应"
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = "基因水平变异影响详情"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    current_row = 4
    for sample in SAMPLES:
        for cat in ("highfreq", "polymorphic"):
            genes_file = annot_dir / cat / f"{sample}.{cat}_snpEff_stats.genes.txt"
            genes = parse_snpeff_genes(genes_file)
            if not genes:
                continue

            for g in genes:
                # 只显示有实际影响的基因
                has_impact = any(
                    int(g.get(f"variants_impact_{imp}", 0)) > 0
                    for imp in ["HIGH", "LOW", "MODERATE", "MODIFIER"]
                )
                if not has_impact:
                    continue

                # 收集存在的效应
                effects = []
                for key, val in g.items():
                    if key.startswith("variants_effect_") and val == "1":
                        effects.append(key.replace("variants_effect_", ""))
                    elif key.startswith("variants_effect_") and val.isdigit() and int(val) > 0:
                        eff_name = key.replace("variants_effect_", "")
                        effects.append(f"{eff_name}({val})")

                ws.cell(row=current_row, column=1, value=sample)
                ws.cell(row=current_row, column=2, value=cat)
                ws.cell(row=current_row, column=3, value=g.get("GeneName", ""))
                ws.cell(row=current_row, column=4, value=g.get("GeneId", ""))
                ws.cell(row=current_row, column=5, value=int(g.get("variants_impact_HIGH", 0)))
                ws.cell(row=current_row, column=6, value=int(g.get("variants_impact_LOW", 0)))
                ws.cell(row=current_row, column=7, value=int(g.get("variants_impact_MODERATE", 0)))
                ws.cell(row=current_row, column=8, value=int(g.get("variants_impact_MODIFIER", 0)))
                ws.cell(row=current_row, column=9, value="; ".join(effects))
                current_row += 1

    style_body(ws, 4, current_row - 1, len(headers))
    auto_width(ws, len(headers))
    ws.sheet_properties.tabColor = "2E75B6"


# ── 主入口 ───────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="生成重测序分析 Excel 报告")
    ap.add_argument("--out", dest="out_path", default=DEFAULT_OUTPUT,
                    help=f"输出 Excel 文件路径 (默认: {DEFAULT_OUTPUT})")
    args = ap.parse_args()

    out_path = Path(args.out_path)
    tables_dir = DEFAULT_TABLES_DIR
    annot_dir = DEFAULT_ANNOT_DIR

    print("📊 正在读取数据...")

    # 汇总信息
    summaries = []
    for s in SAMPLES:
        info = get_sample_summary(s, tables_dir, annot_dir)
        summaries.append(info)
        print(f"  {s}: 高频 {info['highfreq_count']} 个, "
              f"多态 {info['poly_count']} 个")

    print("\n📝 正在生成 Excel 报告...")

    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    create_summary_sheet(wb, summaries)
    print("  ✓ 样本变异汇总")
    create_highfreq_detail_sheet(wb, tables_dir)
    print("  ✓ 高频变异详情")
    create_poly_detail_sheet(wb, tables_dir)
    print("  ✓ 多态变异详情")
    create_comparison_sheet(wb, tables_dir)
    print("  ✓ 变异对比分析")
    create_impact_sheet(wb, annot_dir)
    print("  ✓ snpEff影响分类")
    create_gene_detail_sheet(wb, annot_dir)
    print("  ✓ 基因影响详情")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"\n✅ 报告已生成: {out_path}")
    print(f"   包含 {len(wb.sheetnames)} 个工作表: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
